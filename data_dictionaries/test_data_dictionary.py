"""
Tests for the data dictionary pipeline.

Run with: pytest data_dictionaries/test_data_dictionary.py -v
"""

import json
import subprocess
import sys
from pathlib import Path
from typing import get_type_hints, Optional

import pytest
import yaml
from openpyxl import load_workbook


# Paths
DATA_DIR = Path(__file__).parent
XLSX_PATH = DATA_DIR / '2025_12_ascr_data_dictionary_v1.0.xlsx'
YAML_PATH = DATA_DIR / 'curation_schema.yaml'
PY_PATH = DATA_DIR / 'curation_models.py'
JSONC_PATH = DATA_DIR / 'curation_schema.jsonc'


# --- Test Helpers ---

def load_generated_yaml():
    """Load the generated YAML schema file."""
    with open(YAML_PATH) as f:
        return yaml.safe_load(f)


def load_json_schema():
    """Load the generated JSONC schema file (skips comment lines)."""
    with open(JSONC_PATH) as f:
        lines = [line for line in f if not line.strip().startswith('//')]
        return json.loads(''.join(lines))


def get_unique_class_names_from_xlsx():
    """Extract all unique django_class_name values from source xlsx (excluding PK/FK)."""
    wb = load_workbook(XLSX_PATH)
    ws = wb['data_dictionary']
    class_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        class_name = row[0]
        field_name = row[1]
        key = row[5]
        # Only count if has class name, field name, and is not PK/FK
        if class_name and field_name and (not key or key.upper() not in ('PK', 'FK')):
            class_names.add(class_name)
    return class_names


# --- 1. YAML Generation Tests ---

class TestYAMLGeneration:

    def test_yaml_excludes_pk_fields(self):
        """YAML output should not contain any fields where key='PK'"""
        yaml_data = load_generated_yaml()
        # PK fields are typically 'id' and should be filtered out
        for model_name, model_def in yaml_data.items():
            fields = model_def.get('fields', {})
            # 'id' is the common PK name that should be excluded
            assert 'id' not in fields, f"PK field 'id' found in {model_name}"

    def test_yaml_excludes_fk_fields(self):
        """YAML output should not contain FK fields (verified by checking xlsx)"""
        yaml_data = load_generated_yaml()

        # Get FK field names from xlsx
        wb = load_workbook(XLSX_PATH)
        ws = wb['data_dictionary']
        fk_fields = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[5] and row[5].upper() == 'FK':
                fk_fields.add((row[0], row[1]))  # (class_name, field_name)

        # Verify none of them appear in YAML
        for model_name, model_def in yaml_data.items():
            fields = model_def.get('fields', {})
            for field_name in fields:
                assert (model_name, field_name) not in fk_fields, \
                    f"FK field {model_name}.{field_name} should be excluded"

    def test_yaml_contains_all_models(self):
        """YAML should contain all django_class_name values from xlsx"""
        yaml_data = load_generated_yaml()
        expected_models = get_unique_class_names_from_xlsx()
        assert set(yaml_data.keys()) == expected_models

    def test_yaml_field_metadata_complete(self):
        """Each field should have all required metadata keys"""
        required_keys = [
            'description', 'data_type', 'allows_null', 'field_length',
            'valid_values_long', 'uses_ontology', 'llm_curate', 'llm_instructions'
        ]
        yaml_data = load_generated_yaml()
        for model_name, model_def in yaml_data.items():
            for field_name, field_def in model_def.get('fields', {}).items():
                for key in required_keys:
                    assert key in field_def, f"Missing '{key}' in {model_name}.{field_name}"

    def test_yaml_enum_values_parsed_as_list(self):
        """valid_values_long should be parsed into a list for ENUM fields"""
        yaml_data = load_generated_yaml()
        cell_line = yaml_data['CellLine']['fields']
        cell_type = cell_line['cell_type']
        assert isinstance(cell_type['valid_values_long'], list)
        assert "human embryonic stem cell (hESC)" in cell_type['valid_values_long']

    def test_llm_curate_fields_have_instructions(self):
        """All fields with llm_curate=True should have non-empty instructions"""
        yaml_data = load_generated_yaml()
        missing_instructions = []

        for model_name, model_def in yaml_data.items():
            for field_name, field_def in model_def.get('fields', {}).items():
                if field_def.get('llm_curate'):
                    instruction = field_def.get('llm_instructions')
                    # Check for missing or empty instructions
                    if not instruction:
                        missing_instructions.append(f"{model_name}.{field_name}")

        if missing_instructions:
            pytest.fail(
                f"Fields with llm_curate=True missing instructions:\n" +
                "\n".join(f"  - {f}" for f in missing_instructions)
            )


# --- 2. Pydantic Model Generation Tests ---

class TestPydanticGeneration:

    def test_generated_models_importable(self):
        """Generated curation_models.py should be importable without errors"""
        from data_dictionaries import curation_models
        assert hasattr(curation_models, 'CellLine')

    def test_varchar_maps_to_str(self):
        """VARCHAR fields should be typed as str (or Optional[str])"""
        from data_dictionaries.curation_models import CellLine
        # hpscreg_name is VARCHAR and nullable
        field = CellLine.model_fields['hpscreg_name']
        assert field.annotation == Optional[str] or 'str' in str(field.annotation)

    def test_nullable_fields_are_optional(self):
        """Fields with allows_null=Yes should be Optional with default None"""
        from data_dictionaries.curation_models import CellLine
        # hpscreg_name is nullable
        field = CellLine.model_fields['hpscreg_name']
        assert field.default is None

    def test_required_fields_no_default(self):
        """Fields with allows_null=No should be required (no default)"""
        from data_dictionaries.curation_models import CellLine
        # cell_type is required (allows_null=No)
        field = CellLine.model_fields['cell_type']
        assert field.is_required()

    def test_field_length_constraint(self):
        """String fields should have max_length constraint from field_length"""
        from data_dictionaries.curation_models import CellLine
        field = CellLine.model_fields['hpscreg_name']
        # Check max_length is set to 100
        assert field.metadata is not None or hasattr(field, 'max_length')
        # Verify via JSON schema which exposes constraints
        schema = CellLine.model_json_schema()
        hpscreg_schema = schema['properties']['hpscreg_name']
        # Optional fields have anyOf structure
        if 'anyOf' in hpscreg_schema:
            str_schema = next(s for s in hpscreg_schema['anyOf'] if s.get('type') == 'string')
            assert str_schema.get('maxLength') == 100
        else:
            assert hpscreg_schema.get('maxLength') == 100

    def test_enum_field_validation(self):
        """ENUM fields should reject invalid values"""
        from data_dictionaries.curation_models import CellLine
        from pydantic import ValidationError

        # Get all required fields to construct a valid base
        required_fields = {
            'cell_type': 'human induced pluripotent stem cell (hiPSC)',
            'status': 'Characterised',
            'genotype': 'Patient Control',
            'genotype_locus': 'BRCA1',
            'frozen': True,
            'research_use': True,
            'clinical_use': False,
            'commercial_use': False,
            'certificate_of_pluripotency_characterisation': True,
            'publish': True,
            'registered_with_hpscreg': 'Registered',
            'curation_status': 'Reviewed by Australian Stem Cell Registry'
        }

        # Valid value should work
        valid = CellLine(**required_fields)
        assert valid.cell_type == 'human induced pluripotent stem cell (hiPSC)'

        # Invalid value should raise
        invalid_fields = required_fields.copy()
        invalid_fields['cell_type'] = 'INVALID_TYPE'
        with pytest.raises(ValidationError):
            CellLine(**invalid_fields)


# --- 3. JSON Schema Generation Tests ---

class TestJSONSchemaGeneration:

    def test_json_schema_is_valid_json(self):
        """curation_schema.jsonc should be valid JSON (after stripping comments)"""
        schema = load_json_schema()
        assert isinstance(schema, dict)
        assert '$defs' in schema

    def test_json_schema_contains_all_models(self):
        """JSON schema should have definitions for all models"""
        schema = load_json_schema()
        assert 'CellLine' in schema.get('$defs', {})

    def test_json_schema_enum_values(self):
        """ENUM fields should have enum values in JSON schema"""
        schema = load_json_schema()
        cell_line_schema = schema['$defs']['CellLine']
        cell_type_schema = cell_line_schema['properties']['cell_type']
        # Literal types become enum in JSON schema
        assert 'enum' in cell_type_schema


# --- 4. End-to-End Tests ---

class TestEndToEnd:

    def test_pipeline_runs_successfully(self):
        """Running make_data_dictionary.py should complete without errors"""
        result = subprocess.run(
            [sys.executable, str(DATA_DIR / 'make_data_dictionary.py')],
            capture_output=True,
            cwd=DATA_DIR.parent
        )
        assert result.returncode == 0, f"Pipeline failed: {result.stderr.decode()}"

    def test_all_output_files_created(self):
        """Pipeline should create all three output files"""
        assert YAML_PATH.exists(), "YAML file not created"
        assert PY_PATH.exists(), "Python file not created"
        assert JSONC_PATH.exists(), "JSONC file not created"

    def test_source_file_header_in_yaml(self):
        """YAML should have source file in header comment"""
        with open(YAML_PATH) as f:
            first_line = f.readline()
        assert '2025_12_ascr_data_dictionary_v1.0.xlsx' in first_line

    def test_source_file_header_in_python(self):
        """Python file should have source file in docstring"""
        with open(PY_PATH) as f:
            content = f.read(500)
        assert '2025_12_ascr_data_dictionary_v1.0.xlsx' in content

    def test_source_file_header_in_jsonc(self):
        """JSONC should have source file in header comment"""
        with open(JSONC_PATH) as f:
            first_line = f.readline()
        assert '2025_12_ascr_data_dictionary_v1.0.xlsx' in first_line

    def test_sample_cell_line_validates(self):
        """A sample CellLine JSON should validate against generated model"""
        from data_dictionaries.curation_models import CellLine

        sample = {
            "hpscreg_name": "Genea019",
            "cell_type": "human induced pluripotent stem cell (hiPSC)",
            "status": "Characterised",
            "genotype": "Patient Control",
            "genotype_locus": "BRCA1",
            "frozen": True,
            "research_use": True,
            "clinical_use": False,
            "commercial_use": False,
            "certificate_of_pluripotency_characterisation": True,
            "publish": True,
            "registered_with_hpscreg": "Registered",
            "curation_status": "Reviewed by Australian Stem Cell Registry"
        }
        cell_line = CellLine(**sample)
        assert cell_line.hpscreg_name == "Genea019"
